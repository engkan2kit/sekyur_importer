
import openpyxl
import psycopg2

conn = psycopg2.connect(
    host="localhost",
    database="sekyurlink",
    user="postgres",
    password="root")

curs = conn.cursor()
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("details.xlsx", data_only=True)

work_sheets = dataframe.sheetnames
work_sheets = ['Buenavista Townhomes Clients', 'Biclatan Clients']
# Define variable to read sheet
#dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
#for row in range(0, dataframe1.max_row):
#    for col in dataframe1.iter_cols(1, dataframe1.max_column):
#       print(col[row].value)

for sheet in work_sheets:
    num_rows= 0
    print("Importing from", sheet)
    ws = dataframe[sheet]
    row_iter = 4
    while (True):
        cur_row = ws[row_iter]
        row_iter += 1
        if cur_row[1].value is None:
            if row_iter<=5:
                continue
            else:
                break
        """ for cell in cur_row:
            print(cell.value, end=";")
        print('') 
        """
        try:
            fb_msgrs = list(cur_row[6].value.split())
        except:
            fb_msgrs = []
        
        client_name = cur_row[1].value
        address = cur_row[2].value
        cp1 = cur_row[3].value
        cp2 = cur_row[4].value
        email = cur_row[5].value
        fb_msgr = cur_row[6].value
        id_type = cur_row[7].value
        installation_date = cur_row[8].value

        # We can insert this row at this point.
        curs.execute("INSERT INTO public.subscription (last_name,billing_address, cpnum_1, cpnum_2, email, fb_link,id_type, installation_address,date_created,date_updated,reg_date) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)", (client_name,address,cp1,cp2,email,fb_msgr,id_type,installation_date))
        conn.commit()
        num_rows += 1
    print(num_rows)
curs.close()
conn.close()